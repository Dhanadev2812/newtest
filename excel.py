import openpyxl
path = "C:\\Python27\\PycharmProjects\\Digicampushybrid\\inputdata.xlsx"

def rowCount (File,Sheetname):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(Sheetname)
    return (sheet.max_row)


def columnCount (File,Sheetname):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(Sheetname)
    return (sheet.max_column)


def readData (File,Sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(Sheetname)
    return sheet.cell(row=rownum,column =columnnum).value
    workbook.save(File)


def writeData (File,Sheetname,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(Sheetname)
    sheet.cell(row=rownum,column =columnnum).value = data
    workbook.save(File)


